import os
import sys
import uuid
import csv
import io
import time
import asyncio
import sqlite3
from typing import List, Optional
from pydantic import BaseModel
import pandas as pd

# Set event loop policy on Windows for Playwright subprocess support
if sys.platform == 'win32':
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

from fastapi import FastAPI, UploadFile, File, Form, BackgroundTasks, HTTPException, Request
from fastapi.responses import FileResponse, StreamingResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from services.tracking_service import TrackingService
from scrapers.factory import ScraperFactory

app = FastAPI(title="TrackShip API")

# Allow CORS for development ease
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "static")
SCREENSHOTS_DIR = os.path.join(STATIC_DIR, "screenshots")
UPLOADS_DIR = os.path.join(BASE_DIR, "uploads")
OUTPUTS_DIR = os.path.join(BASE_DIR, "outputs")
DB_PATH = os.path.join(BASE_DIR, "tracking.db")

# Ensure folders exist
os.makedirs(UPLOADS_DIR, exist_ok=True)
os.makedirs(OUTPUTS_DIR, exist_ok=True)
os.makedirs(SCREENSHOTS_DIR, exist_ok=True)

def get_courier_direct_url(courier: str, awb: str) -> str:
    c = (courier or "").lower()
    awb_upper = (awb or "").upper().strip()
    if "shadowfax" in c or awb_upper.startswith("SF") or awb_upper.startswith("R"):
        return f"https://trackcourier.io/track-and-trace/shadowfax/{awb}"
    elif "delhivery" in c:
        return f"https://www.delhivery.com/track/package/{awb}"
    elif "bluedart" in c or "blue dart" in c:
        return f"https://www.bluedart.com/tracking"
    elif "ekart" in c or awb_upper.startswith("FMP") or awb_upper.startswith("EKART"):
        return f"https://ekartlogistics.com/"
    elif "xpressbees" in c:
        return f"https://www.xpressbees.com/track?isawb=Yes&trackid={awb}"
    elif "dtdc" in c:
        return f"https://www.dtdc.in/tracking.asp"
    elif "ecom" in c:
        return f"https://ecomexpress.in/tracking/?awb_field={awb}"
    elif "india post" in c or "indiapost" in c:
        return f"https://www.indiapost.gov.in/_layouts/15/dop.portal.tracking/trackconsignment.aspx"
    else:
        return f"https://www.delhivery.com/track/package/{awb}"

@app.get('/static/screenshots/{filename}')
async def serve_screenshot(filename: str):
    """
    Intelligent screenshot server:
    1. Returns cached screenshot if present on disk.
    2. If missing (e.g. Render server restarted/slept), tries live-capturing screenshot.
    3. If capture fails or times out, seamlessly redirects to live official courier tracking page.
    Never returns 404!
    """
    file_path = os.path.join(SCREENSHOTS_DIR, filename)
    if os.path.exists(file_path):
        return FileResponse(file_path, media_type="image/png")
    
    awb = filename.replace(".png", "").strip()
    
    # Identify courier
    courier = ""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT courier FROM shipments WHERE tracking_number = ? LIMIT 1", (awb,))
        row = cursor.fetchone()
        if row:
            courier = row[0]
        conn.close()
    except Exception:
        pass
        
    if not courier:
        if awb.upper().startswith("R") or awb.upper().startswith("SF"):
            courier = "Shadowfax"
        elif awb.isdigit() and len(awb) in [12, 13, 14, 15]:
            courier = "Delhivery"
        elif awb.upper().startswith("X") or awb.upper().startswith("14"):
            courier = "Xpressbees"
        elif awb.upper().startswith("FMP") or awb.upper().startswith("EKART"):
            courier = "Ekart"
        else:
            courier = "Delhivery"

    # Attempt fast on-demand capture
    scraper = ScraperFactory.get_scraper(courier)
    if scraper:
        try:
            await asyncio.wait_for(scraper.track(awb), timeout=8.0)
            if os.path.exists(file_path):
                return FileResponse(file_path, media_type="image/png")
        except Exception as e:
            print(f"On-demand screenshot capture failed for {awb}: {e}")

    # Fallback to official tracking URL so clicking link in Excel always works
    direct_url = get_courier_direct_url(courier, awb)
    return RedirectResponse(url=direct_url, status_code=307)

# Mount static folder
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON;")
    cursor = conn.cursor()
    # Create tasks table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS tasks (
        task_id TEXT PRIMARY KEY,
        status TEXT,
        progress INTEGER,
        current_action TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
    """)
    # Create shipments table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS shipments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        task_id TEXT,
        invoice_no TEXT,
        tracking_number TEXT,
        courier TEXT,
        platform_status TEXT,
        status TEXT,
        last_location TEXT,
        timestamp TEXT,
        last_sync TEXT,
        screenshot TEXT,
        FOREIGN KEY(task_id) REFERENCES tasks(task_id) ON DELETE CASCADE
    );
    """)
    # Migration step to add last_sync if table already exists
    try:
        cursor.execute("ALTER TABLE shipments ADD COLUMN last_sync TEXT;")
    except sqlite3.OperationalError:
        pass
    # Migration step to add invoice_no if table already exists
    try:
        cursor.execute("ALTER TABLE shipments ADD COLUMN invoice_no TEXT;")
    except sqlite3.OperationalError:
        pass
    # Migration step to add screenshot if table already exists
    try:
        cursor.execute("ALTER TABLE shipments ADD COLUMN screenshot TEXT;")
    except sqlite3.OperationalError:
        pass
    # Migration step to add platform_status if table already exists
    try:
        cursor.execute("ALTER TABLE shipments ADD COLUMN platform_status TEXT;")
    except sqlite3.OperationalError:
        pass
    # Create logs table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        task_id TEXT,
        message TEXT,
        level TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(task_id) REFERENCES tasks(task_id) ON DELETE CASCADE
    );
    """)
    # Create api_usage table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS api_usage (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
    """)
    # Prune task data older than 24 hours
    cursor.execute("DELETE FROM tasks WHERE created_at < datetime('now', '-24 hours');")
    conn.commit()
    conn.close()

# Initialize DB on startup
init_db()

class StartTrackRequest(BaseModel):
    task_id: str
    shipments: Optional[List[dict]] = None

class SyncSingleRequest(BaseModel):
    task_id: str
    tracking_number: str
    courier: str
    invoice_no: Optional[str] = ""
    platform_status: Optional[str] = ""

class RestoreTaskRequest(BaseModel):
    task_id: str
    shipments: List[dict]

class ExportDirectRequest(BaseModel):
    task_id: Optional[str] = "export"
    shipments: List[dict]

@app.get('/')
def root():
    return FileResponse("static/index.html")

def clean_tracking_number(awb_val) -> str:
    if pd.isna(awb_val):
        return ""
    s = str(awb_val).strip()
    if not s:
        return ""
        
    # If string contains scientific notation (e.g. 1.95e+14)
    if 'e' in s.lower():
        try:
            val = float(s)
            return str(int(val)) if val.is_integer() else f"{val:.0f}"
        except ValueError:
            pass
            
    # If float-like (e.g. 195042600200336.0)
    try:
        val = float(s)
        if val.is_integer():
            return str(int(val))
    except ValueError:
        pass
        
    return s

def find_col_value(row, aliases) -> str:
    # Normalized search over row keys
    row_dict = {str(k).strip().lower(): v for k, v in row.items()}
    for alias in aliases:
        norm_alias = alias.strip().lower()
        if norm_alias in row_dict:
            val = row_dict[norm_alias]
            if pd.notna(val):
                return str(val).strip()
    return ""

@app.post('/api/upload')
async def upload_file(file: UploadFile = File(...)):
    filename = file.filename
    ext = os.path.splitext(filename)[1].lower()
    
    if ext not in ['.csv', '.xlsx', '.xls']:
        raise HTTPException(status_code=400, detail="Unsupported file format. Please upload CSV or Excel.")

    contents = await file.read()
    shipments = []

    invoice_aliases = ['invoice no', 'invoice_no', 'invoice no.', 'invoice', 'invoice number', 'inv no', 'inv_no', 'invoice#', 'inv']
    awb_aliases = ['awb', 'awb no', 'awb no.', 'awb number', 'tracking number', 'tracking_number', 'tracking no', 'tracking_no', 'tracking #', 'waybill']
    courier_aliases = ['courier', 'courier partner', 'courier_partner', 'courier name', 'courier_name', 'partner', 'logistic', 'logistics']
    platform_status_aliases = ['platform status', 'platform_status', 'order status', 'order_status', 'platform status name', 'platform_status_name']

    try:
        if ext == '.csv':
            decoded = contents.decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(decoded))
            for row in csv_reader:
                invoice = find_col_value(row, invoice_aliases)
                awb = find_col_value(row, awb_aliases)
                courier = find_col_value(row, courier_aliases)
                platform_status = find_col_value(row, platform_status_aliases)
                
                if awb:
                    clean_awb = clean_tracking_number(awb)
                    if clean_awb:
                        shipments.append({
                            "invoice_no": invoice,
                            "tracking_number": clean_awb,
                            "courier": courier if courier else "Delhivery",
                            "platform_status": platform_status,
                            "status": "Pending",
                            "last_location": "Awaiting scan",
                            "timestamp": "-",
                            "last_sync": "-",
                            "screenshot": "-"
                        })
        else:
            df = pd.read_excel(io.BytesIO(contents), dtype=str)
            for _, row in df.iterrows():
                # row can be converted to dict to use find_col_value
                row_dict = row.to_dict()
                invoice = find_col_value(row_dict, invoice_aliases)
                awb = find_col_value(row_dict, awb_aliases)
                courier = find_col_value(row_dict, courier_aliases)
                platform_status = find_col_value(row_dict, platform_status_aliases)
                
                if awb and str(awb).strip():
                    clean_awb = clean_tracking_number(awb)
                    shipments.append({
                        "invoice_no": invoice,
                        "tracking_number": clean_awb,
                        "courier": courier if courier else "Delhivery",
                        "platform_status": platform_status,
                        "status": "Pending",
                        "last_location": "Awaiting scan",
                        "timestamp": "-",
                        "last_sync": "-",
                        "screenshot": "-"
                        })

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error parsing tracking sheet: {str(e)}")

    if not shipments:
        raise HTTPException(status_code=400, detail="No tracking numbers found in the uploaded sheet. Please check headers (AWB, Courier).")

    task_id = str(uuid.uuid4())
    
    # Save upload to database
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("INSERT INTO tasks (task_id, status, progress, current_action) VALUES (?, ?, ?, ?)", (task_id, "pending", 0, "Ready to start"))
    
    for s in shipments:
        cursor.execute("""
        INSERT INTO shipments (task_id, invoice_no, tracking_number, courier, platform_status, status, last_location, timestamp, last_sync, screenshot)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (task_id, s.get("invoice_no", ""), s["tracking_number"], s["courier"], s.get("platform_status", ""), s["status"], s["last_location"], s["timestamp"], "-", "-"))
        
    cursor.execute("INSERT INTO logs (task_id, message, level) VALUES (?, ?, ?)", (task_id, f"Successfully parsed {filename}. Found {len(shipments)} records.", "success"))
    cursor.execute("INSERT INTO logs (task_id, message, level) VALUES (?, ?, ?)", (task_id, "Ready to begin courier web scraping simulation.", "info"))
    
    # Get today's API calls count
    cursor.execute("SELECT COUNT(*) FROM api_usage WHERE timestamp >= datetime('now', 'start of day');")
    today_api_calls = cursor.fetchone()[0]

    conn.commit()
    conn.close()

    # Calculate stats
    delivered = sum(1 for s in shipments if s["status"].lower() == "delivered")
    transit = sum(1 for s in shipments if s["status"].lower() in ["in transit", "out for delivery", "picked up", "out for pickup"])
    failed = sum(1 for s in shipments if "failed" in s["status"].lower() or "invalid" in s["status"].lower() or "error" in s["status"].lower())

    stats = {
        "total": len(shipments),
        "delivered": delivered,
        "transit": transit,
        "failed": failed,
        "api_calls": today_api_calls
    }

    return {"task_id": task_id, "shipments": shipments, "stats": stats}


# Real background task runner calling TrackingService
async def run_tracking_simulation(task_id: str):
    # Retrieve shipments from database
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT tracking_number, courier, status, last_location, timestamp, last_sync, screenshot FROM shipments WHERE task_id = ?", (task_id,))
    rows = cursor.fetchall()
    conn.close()
    
    shipments = []
    for r in rows:
        shipments.append({
            "tracking_number": r[0],
            "courier": r[1],
            "status": r[2],
            "last_location": r[3],
            "timestamp": r[4],
            "last_sync": r[5] or "-",
            "screenshot": r[6] or "-"
        })
        
    # Set status to running
    conn = sqlite3.connect(DB_PATH)
    conn.execute("UPDATE tasks SET status = ? WHERE task_id = ?", ("running", task_id))
    conn.commit()
    conn.close()
    
    # We define progress callback to update SQLite task
    async def progress_callback(progress, current_action, log_message, log_level):
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("UPDATE tasks SET progress = ?, current_action = ? WHERE task_id = ?", (progress, current_action, task_id))
        
        # Update shipments status in database
        for s in shipments:
            cursor.execute("""
            UPDATE shipments 
            SET status = ?, last_location = ?, timestamp = ?, last_sync = ?, screenshot = ? 
            WHERE task_id = ? AND tracking_number = ?
            """, (s["status"], s["last_location"], s["timestamp"], s.get("last_sync", "-"), s.get("screenshot", "-"), task_id, s["tracking_number"]))
            
        cursor.execute("INSERT INTO logs (task_id, message, level) VALUES (?, ?, ?)", (task_id, log_message, log_level))
        conn.commit()
        conn.close()

    try:
        await TrackingService.track_shipments(shipments, task_id, progress_callback)
        conn = sqlite3.connect(DB_PATH)
        conn.execute("UPDATE tasks SET status = ? WHERE task_id = ?", ("completed", task_id))
        conn.commit()
        conn.close()
    except Exception as e:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("UPDATE tasks SET status = ? WHERE task_id = ?", ("failed", task_id))
        cursor.execute("INSERT INTO logs (task_id, message, level) VALUES (?, ?, ?)", (task_id, f"Fatal tracking engine error: {str(e)}", "error"))
        conn.commit()
        conn.close()


class QuerySingleRequest(BaseModel):
    tracking_number: str
    courier: str

@app.post('/api/track/query_single')
async def query_single_shipment(body: QuerySingleRequest):
    awb = body.tracking_number.strip()
    courier = body.courier.strip()
    
    if not awb or not courier:
        raise HTTPException(status_code=400, detail="AWB number and Courier are required")
        
    scraper = ScraperFactory.get_scraper(courier)
    if not scraper:
        raise HTTPException(status_code=400, detail=f"Courier '{courier}' not supported")
        
    # Track immediately
    try:
        result = await scraper.track(awb)
        status = result["status"]
        last_location = result["last_location"]
        timestamp = result["timestamp"]
        screenshot = result.get("screenshot", "-")
    except Exception as e:
        status = "Scrape Error"
        last_location = f"Error: {str(e)}"
        timestamp = "-"
        screenshot = "-"
        
    # API calls tracking
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("INSERT INTO api_usage DEFAULT VALUES;")
    cursor.execute("SELECT COUNT(*) FROM api_usage WHERE timestamp >= datetime('now', 'start of day');")
    today_api_calls = cursor.fetchone()[0]
    conn.commit()
    conn.close()
    
    return {
        "tracking_number": awb,
        "courier": courier,
        "status": status,
        "last_location": last_location,
        "timestamp": timestamp,
        "screenshot": screenshot,
        "api_calls": today_api_calls
    }

@app.post('/api/track/start')
async def start_tracking(body: StartTrackRequest, background_tasks: BackgroundTasks):
    task_id = body.task_id
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM tasks WHERE task_id = ?", (task_id,))
    exists = cursor.fetchone()
    
    # Self-healing: If task doesn't exist in DB (e.g. server woke up or restarted), re-create it if shipments are sent
    if not exists:
        if body.shipments:
            cursor.execute("INSERT OR REPLACE INTO tasks (task_id, status, progress, current_action) VALUES (?, ?, ?, ?)", (task_id, "pending", 0, "Ready to start"))
            for s in body.shipments:
                cursor.execute("""
                INSERT INTO shipments (task_id, invoice_no, tracking_number, courier, platform_status, status, last_location, timestamp, last_sync, screenshot)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    task_id,
                    s.get("invoice_no", ""),
                    s["tracking_number"],
                    s.get("courier", "Delhivery"),
                    s.get("platform_status", ""),
                    s.get("status", "Pending"),
                    s.get("last_location", "Awaiting scan"),
                    s.get("timestamp", "-"),
                    s.get("last_sync", "-"),
                    s.get("screenshot", "-")
                ))
            cursor.execute("INSERT INTO logs (task_id, message, level) VALUES (?, ?, ?)", (task_id, f"Restored {len(body.shipments)} records into task session.", "info"))
            conn.commit()
        else:
            conn.close()
            raise HTTPException(status_code=404, detail="Task ID not found")
            
    conn.close()
    background_tasks.add_task(run_tracking_simulation, task_id)
    return {"status": "started"}


@app.post('/api/track/sync_single')
async def sync_single_shipment(body: SyncSingleRequest):
    task_id = body.task_id
    awb = body.tracking_number
    courier = body.courier
    
    scraper = ScraperFactory.get_scraper(courier)
    if not scraper:
        raise HTTPException(status_code=400, detail=f"Courier '{courier}' not supported")
        
    try:
        result = await scraper.track(awb)
        
        from datetime import datetime
        last_sync_str = datetime.now().strftime("%d-%m-%Y %I:%M:%S %p")
        
        # Self-healing: Ensure task and shipment exist in DB
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("INSERT INTO api_usage DEFAULT VALUES;")
        cursor.execute("INSERT OR IGNORE INTO tasks (task_id, status, progress, current_action) VALUES (?, ?, ?, ?)", (task_id, "completed", 100, "Idle"))
        
        cursor.execute("SELECT 1 FROM shipments WHERE task_id = ? AND tracking_number = ?", (task_id, awb))
        if not cursor.fetchone():
            cursor.execute("""
            INSERT INTO shipments (task_id, invoice_no, tracking_number, courier, platform_status, status, last_location, timestamp, last_sync, screenshot)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (task_id, body.invoice_no or "", awb, courier, body.platform_status or "", result.get("status"), result.get("last_location"), result.get("timestamp"), last_sync_str, result.get("screenshot", "-")))
        else:
            cursor.execute("""
            UPDATE shipments 
            SET status = ?, last_location = ?, timestamp = ?, last_sync = ?, screenshot = ? 
            WHERE task_id = ? AND tracking_number = ?
            """, (result.get("status"), result.get("last_location"), result.get("timestamp"), last_sync_str, result.get("screenshot", "-"), task_id, awb))
        
        # Log the manual update
        cursor.execute("""
        INSERT INTO logs (task_id, message, level) 
        VALUES (?, ?, ?)
        """, (task_id, f"Manually synced {courier} AWB {awb}. Status: {result.get('status')}", "success"))
        
        # Get today's API calls count
        cursor.execute("SELECT COUNT(*) FROM api_usage WHERE timestamp >= datetime('now', 'start of day');")
        today_api_calls = cursor.fetchone()[0]

        conn.commit()
        conn.close()
        
        return {
            "status": result.get("status"),
            "last_location": result.get("last_location"),
            "timestamp": result.get("timestamp"),
            "last_sync": last_sync_str,
            "screenshot": result.get("screenshot", "-"),
            "api_calls": today_api_calls
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Scrape error: {str(e)}")


@app.get('/api/track/progress')
async def get_progress(task_id: str):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT status, progress, current_action FROM tasks WHERE task_id = ?", (task_id,))
    task_row = cursor.fetchone()
    if not task_row:
        conn.close()
        raise HTTPException(status_code=404, detail="Task ID not found")
        
    status, progress, current_action = task_row
    
    # Get shipments
    cursor.execute("SELECT invoice_no, tracking_number, courier, platform_status, status, last_location, timestamp, last_sync, screenshot FROM shipments WHERE task_id = ?", (task_id,))
    shipment_rows = cursor.fetchall()
    shipments = []
    for r in shipment_rows:
        shipments.append({
            "invoice_no": r[0] or "",
            "tracking_number": r[1],
            "courier": r[2],
            "platform_status": r[3] or "",
            "status": r[4],
            "last_location": r[5],
            "timestamp": r[6],
            "last_sync": r[7] or "-",
            "screenshot": r[8] or "-"
        })
        
    # Get logs
    cursor.execute("SELECT message, level FROM logs WHERE task_id = ? ORDER BY id ASC", (task_id,))
    log_rows = cursor.fetchall()
    logs = [{"message": r[0], "level": r[1]} for r in log_rows]
    
    # Clear logs for this poll (so they are only displayed once on console)
    cursor.execute("DELETE FROM logs WHERE task_id = ?", (task_id,))
    
    # Get today's API calls count
    cursor.execute("SELECT COUNT(*) FROM api_usage WHERE timestamp >= datetime('now', 'start of day');")
    today_api_calls = cursor.fetchone()[0]
    
    conn.commit()
    conn.close()
    
    # Calculate stats
    delivered = sum(1 for s in shipments if s["status"].lower() == "delivered")
    transit = sum(1 for s in shipments if s["status"].lower() in ["in transit", "out for delivery", "picked up", "out for pickup"])
    failed = sum(1 for s in shipments if "failed" in s["status"].lower() or "invalid" in s["status"].lower() or "error" in s["status"].lower())
    
    stats = {
        "total": len(shipments),
        "delivered": delivered,
        "transit": transit,
        "failed": failed,
        "api_calls": today_api_calls
    }
    
    return {
        "status": status,
        "progress": progress,
        "current_action": current_action,
        "shipments": shipments,
        "stats": stats,
        "logs": logs
    }


def generate_excel_stream(shipment_rows, task_id: str, request: Request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tracking Results"

    # Define headers
    headers = ["Invoice No.", "AWB No.", "Courier Partner", "Platform Status", "Status", "Last Location", "Timestamp", "Last Sync", "Screenshot"]
    ws.append(headers)

    # Set Header styling (bold, light gray background, center align)
    header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="334155")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Add data rows with colors matching the frontend palette
    AWB_COLORS_HEX = [
        '1E40AF', '9D174D', '065F46', '92400E', '5B21B6',
        'C2410C', '155E75', '9F1239', '166534', '3730A3',
        '854D0E', '6B21A8', '134E4A', '991B1B', '075985',
        '86198F', '3F6212', '334155', 'BE123C', '115E59',
        'A21CAF', '14532D', '1E3A8A', '78350F', '831843',
        '4C1D95', '064E3B', '9A3412', '0F172A', '713F12'
    ]

    for idx, r in enumerate(shipment_rows):
        row_num = idx + 2
        row_color = AWB_COLORS_HEX[idx % len(AWB_COLORS_HEX)]
        row_font = Font(name="Segoe UI", size=11, color=row_color)
        
        # Write values
        values = [
            r[0] or "",
            r[1] or "",
            r[2] or "",
            r[3] or "",
            r[4] or "",
            r[5] or "",
            r[6] or "",
            r[7] or "-"
        ]
        
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = row_font
            # Alignments: Left align for text/location, Center for status/numbers/dates
            if col_idx in [1, 2, 4, 5, 7, 8]: # Invoice, AWB, Platform Status, Status, Timestamp, Last Sync
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")
                
        # Write Screenshot hyperlink column (col_idx = 9)
        screenshot_path = r[8]
        cell = ws.cell(row=row_num, column=9)
        if screenshot_path and screenshot_path != "-":
            # Construct absolute URL
            base_url_str = str(request.base_url).rstrip('/')
            if str(screenshot_path).startswith("http"):
                screenshot_url = screenshot_path
            else:
                screenshot_url = f"{base_url_str}{screenshot_path}"
            cell.value = "Link"
            cell.hyperlink = screenshot_url
            cell.font = Font(name="Segoe UI", size=11, color="0000FF", underline="single")
        else:
            cell.value = "-"
            cell.font = row_font
        cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Save to dynamic buffer
    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)

    filename_suffix = (task_id[:8] if task_id else "export")
    return StreamingResponse(
        out_buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=tracking_export_{filename_suffix}.xlsx"}
    )


@app.get('/api/export')
async def export_results(task_id: str, request: Request):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT invoice_no, tracking_number, courier, platform_status, status, last_location, timestamp, last_sync, screenshot FROM shipments WHERE task_id = ?", (task_id,))
    shipment_rows = cursor.fetchall()
    conn.close()
    
    if not shipment_rows:
        raise HTTPException(status_code=404, detail="Task ID not found")
        
    return generate_excel_stream(shipment_rows, task_id, request)


@app.post('/api/export_direct')
async def export_direct(body: ExportDirectRequest, request: Request):
    """Export Excel directly from client-supplied shipments if server DB was reset."""
    rows = []
    for s in body.shipments:
        rows.append((
            s.get("invoice_no", ""),
            s.get("tracking_number", ""),
            s.get("courier", ""),
            s.get("platform_status", ""),
            s.get("status", ""),
            s.get("last_location", ""),
            s.get("timestamp", ""),
            s.get("last_sync", "-"),
            s.get("screenshot", "-")
        ))
    return generate_excel_stream(rows, body.task_id or "export", request)


@app.post('/api/restore_task')
async def restore_task(body: RestoreTaskRequest):
    """Restore task and shipments into SQLite database from client session storage."""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("INSERT OR REPLACE INTO tasks (task_id, status, progress, current_action) VALUES (?, ?, ?, ?)", (body.task_id, "completed", 100, "Restored from session"))
    cursor.execute("DELETE FROM shipments WHERE task_id = ?", (body.task_id,))
    for s in body.shipments:
        cursor.execute("""
        INSERT INTO shipments (task_id, invoice_no, tracking_number, courier, platform_status, status, last_location, timestamp, last_sync, screenshot)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            body.task_id,
            s.get("invoice_no", ""),
            s["tracking_number"],
            s.get("courier", "Delhivery"),
            s.get("platform_status", ""),
            s.get("status", "Pending"),
            s.get("last_location", "Awaiting scan"),
            s.get("timestamp", "-"),
            s.get("last_sync", "-"),
            s.get("screenshot", "-")
        ))
    conn.commit()
    conn.close()
    return {"status": "restored", "count": len(body.shipments)}



@app.get('/api/latest')
async def get_latest_task():
    """Return the most recent task's shipments so the UI can restore on page refresh."""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # Get the most recent task
    cursor.execute("SELECT task_id FROM tasks ORDER BY created_at DESC LIMIT 1")
    task_row = cursor.fetchone()

    if not task_row:
        conn.close()
        return {"task_id": None, "shipments": [], "stats": {"total": 0, "delivered": 0, "transit": 0, "failed": 0, "api_calls": 0}}

    task_id = task_row[0]

    # Get shipments
    cursor.execute("SELECT invoice_no, tracking_number, courier, platform_status, status, last_location, timestamp, last_sync, screenshot FROM shipments WHERE task_id = ?", (task_id,))
    shipment_rows = cursor.fetchall()
    shipments = []
    for r in shipment_rows:
        shipments.append({
            "invoice_no": r[0] or "",
            "tracking_number": r[1],
            "courier": r[2],
            "platform_status": r[3] or "",
            "status": r[4],
            "last_location": r[5],
            "timestamp": r[6],
            "last_sync": r[7] or "-",
            "screenshot": r[8] or "-"
        })

    # Get today's API calls count
    cursor.execute("SELECT COUNT(*) FROM api_usage WHERE timestamp >= datetime('now', 'start of day');")
    today_api_calls = cursor.fetchone()[0]

    conn.close()

    # Calculate stats
    delivered = sum(1 for s in shipments if s["status"].lower() == "delivered")
    transit = sum(1 for s in shipments if s["status"].lower() in ["in transit", "out for delivery", "picked up", "out for pickup"])
    failed = sum(1 for s in shipments if "failed" in s["status"].lower() or "invalid" in s["status"].lower() or "error" in s["status"].lower())

    stats = {
        "total": len(shipments),
        "delivered": delivered,
        "transit": transit,
        "failed": failed,
        "api_calls": today_api_calls
    }

    return {"task_id": task_id, "shipments": shipments, "stats": stats}


@app.delete('/api/clear')
async def clear_all_data():
    """Delete all tasks, shipments, and logs from the database."""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM logs")
    cursor.execute("DELETE FROM shipments")
    cursor.execute("DELETE FROM tasks")
    conn.commit()
    conn.close()
    return {"status": "cleared"}
